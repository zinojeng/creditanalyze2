import streamlit as st
import subprocess
import sys

# Add this at the beginning of your script
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Check if cv2 is installed, if not, install it
try:
    import cv2
except ImportError:
    st.info("Installing required packages...")
    install("opencv-python-headless")
    import cv2

# Rest of your imports
import openai
from openai import OpenAI
import os
import json
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import textract
import base64
import tempfile
from dotenv import load_dotenv
from PIL import Image
import pytesseract
import PyPDF2
import numpy as np
from pdf2image import convert_from_path
import re
import logging
import textwrap
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF

load_dotenv()  # 載入 .env 檔案中的環境變數

# 設置日誌
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

VERSION = "1.0.3"

def preprocess_image(image):
    # 轉換為灰度圖
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # 應用自適應閾值
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    
    # 去噪
    denoised = cv2.fastNlMeansDenoising(thresh, None, 10, 7, 21)
    
    return Image.fromarray(denoised)

def ocr_image(image):
    preprocessed_image = preprocess_image(image)
    text = pytesseract.image_to_string(preprocessed_image, lang='chi_tra+eng')
    return text

def extract_text_from_pdf_pypdf2(file_path):
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        logger.error(f"PyPDF2 從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf_ocr(file_path):
    try:
        images = convert_from_path(file_path)
        text = ""
        for image in images:
            text += ocr_image(image) + "\n\n"
        return text
    except Exception as e:
        logger.error(f"OCR 從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf_gpt4_vision(client, file_path, model="gpt-4.1-mini"):
    try:
        images = convert_from_path(file_path)
        all_text = ""
        for i, image in enumerate(images):
            # 保存圖片到臨時文件
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_file:
                image.save(temp_file, format="PNG")
                temp_file_path = temp_file.name

            # 讀取圖片文件
            with open(temp_file_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')

            response = client.chat.completions.create(
                model=model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "請閱讀這張圖片，並提取所有可見的文字內容。只需返回提取的文字，不需要任何解釋或格式化。"},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            },
                        ],
                    }
                ],
                max_tokens=1000,
            )
            
            all_text += response.choices[0].message.content + "\n\n"
            
            # 刪除臨時文件
            os.unlink(temp_file_path)

        return all_text.strip()
    except Exception as e:
        logger.error(f"Vision 模型從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf(client, file_path, model="gpt-4.1-mini"):
    # 首先嘗試使用 PyMuPDF
    try:
        doc = fitz.open(file_path)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    except Exception as e:
        logger.error(f"PyMuPDF 從 PDF 提取文本時出錯: {str(e)}")
        
    # 如果 PyMuPDF 提取失敗，則使用 Vision 模型
    logger.info(f"PyMuPDF 提取失敗，切換到 {model}")
    return extract_text_from_pdf_gpt4_vision(client, file_path, model)

def clean_text(text):
    # 移除多餘的空白字符
    text = re.sub(r'\s+', ' ', text)
    # 移除非打印字符
    text = ''.join(char for char in text if char.isprintable() or char.isspace())
    return text.strip()

def analyze_file(client, file_path, model="gpt-4.1-mini"):
    file_extension = os.path.splitext(file_path)[1].lower()
    try:
        if file_extension == '.pdf':
            text = extract_text_from_pdf(client, file_path, model)  # 傳入模型參數
        elif file_extension in ['.doc', '.docx']:
            text = textract.process(file_path).decode('utf-8', errors='ignore')
        elif file_extension in ['.jpg', '.jpeg', '.png']:
            image = Image.open(file_path)
            text = ocr_image(image)
        else:
            logger.error(f"不支援的檔案類型：{file_extension}")
            return None

        text = clean_text(text)
        
        if not text.strip():
            logger.warning(f"無法從檔案 {file_path} 提取文本")
            return None

        return text
    except Exception as e:
        logger.error(f"處理檔案 {file_path} 時出錯: {str(e)}")
        return None

def calculate_credits(duration_minutes, credit_type):
    if credit_type == "甲類":
        return math.floor(duration_minutes / 50)
    else:  # 乙類
        full_credits = duration_minutes // 50
        remaining_minutes = duration_minutes % 50
        if remaining_minutes >= 25:  # 如果剩餘分鐘數超過或等於 25，則四捨五入
            full_credits += 1
        return round(full_credits * 0.5, 1)  # 乙類每學分為 0.5 分

def get_gpt4_json_response(client, prompt, model="gpt-4.1-mini"):
    try:
        # 判斷是否為 o 系列或 gpt-4o
        is_o_series = model.startswith("o3-") or model.startswith("o4-") or model == "gpt-4o"
        
        params = {
            "model": model,
            "messages": [
                {
                    "role": "system", 
                    "content": """你是一個專業的醫學會議文件分析助手。你的任務是從會議文件中提取關鍵資訊並以 JSON 格式返回。

請遵循以下規則：
1. 仔細閱讀整個文件內容
2. 識別並提取所有必要的欄位
3. 確保所有時間格式統一（HH:MM）
4. 如果無法確定某個欄位的值，使用空字串而不是猜測
5. 對於 AI 初審欄位，嚴格按照規則判斷相關性
6. 確保輸出的 JSON 格式完全符合指定的結構

你的分析必須準確且一致。如果遇到模糊或不確定的情況，寧可留空也不要猜測。"""
                },
                {"role": "user", "content": prompt}
            ],
            "response_format": {"type": "json_object"}
        }
        
        # 根據模型類型添加不同的 tokens 參數
        if is_o_series:
            params["max_completion_tokens"] = 1500
        else:
            params["max_tokens"] = 1500
            
        response = client.chat.completions.create(**params)
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        logger.error(f"OpenAI API 調用失敗: {str(e)}")
        return None

def calculate_duration(start_time, end_time):
    start = datetime.strptime(start_time, "%H:%M")
    end = datetime.strptime(end_time, "%H:%M")
    if end < start:
        end += timedelta(days=1)  # 處理跨午夜的情況
    duration = end - start
    return int(duration.total_seconds() / 60)  # 返回分鐘數

def process_topics(topics):
    processed_topics = []
    for topic in topics:
        processed_topic = {
            "topic": topic.get("topic", ""),
            "speaker": topic.get("speaker", ""),  # 保留為空字串如果沒有講者
            "moderator": topic.get("moderator", ""),
            "time": topic.get("time", ""),
            "duration": int(topic.get("duration", 0)),
            "ai_review": topic.get("ai_review", "")
        }
        processed_topics.append(processed_topic)
    return processed_topics

def write_to_excel(all_results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "課程分析結果"

    # 設置標題樣式
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    # 設置文件名樣式
    filename_font = Font(bold=True, size=14)  # 增加字體大小
    filename_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # 寫入標題
    headers = ["文件名", "主題", "主辦單位", "日期", "地點", "積分類別", "原積分數", "AI初審積分"]
    column_widths = [20, 65, 45, 25, 15, 10, 10, 10]  # 為每列設置寬度

    for col, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.alignment = title_alignment
        ws.column_dimensions[get_column_letter(col)].width = width

    # 寫入每個文件的基本信息
    for row, result in enumerate(all_results, start=2):
        ws.cell(row=row, column=1, value=result.get('文件名', ''))
        ws.cell(row=row, column=2, value=result.get('主題', ''))
        ws.cell(row=row, column=3, value=result.get('主辦單位', ''))
        ws.cell(row=row, column=4, value=result.get('日期', ''))
        ws.cell(row=row, column=5, value=result.get('地點', ''))
        ws.cell(row=row, column=6, value=result.get('積分類別', ''))
        ws.cell(row=row, column=7, value=result.get('原始積分數', 0))
        ws.cell(row=row, column=8, value=result.get('AI初審積分', 0))

    # 創建詳細分析結果工作表
    ws_detail = wb.create_sheet(title="詳細分析結果")
    
    current_row = 1
    for result in all_results:
        # 寫入文件名作為分隔，並設置淡綠色背景延伸到最後一欄
        filename_text = f"文件名: {result.get('文件名', '')}"
        
        # 將淡綠色背景延伸到最後一欄（F欄）
        for col in range(1, 7):  # 從第1欄到第6欄（F欄）
            cell = ws_detail.cell(row=current_row, column=col)
            cell.fill = filename_fill
            if col == 1:
                cell.value = filename_text
                cell.font = filename_font
        
        current_row += 1

        # 寫入基本信息
        basic_info = [
            ("主題", result.get('主題', '')),
            ("主辦單位", result.get('主辦單位', '')),
            ("日期", result.get('日期', '')),
            ("地點", result.get('地點', '')),
            ("積分類別", result.get('積分類別', '')),
            ("原始積分數", result.get('原始積分數', 0)),
            ("AI初審積分", result.get('AI初審積分', 0)),
            ("AI初審積分說明", result.get('AI初審積分說明', '')),
            ("AI初審詳細說明", result.get('AI初審詳細說明', '')),
            ("積分計算方式", result.get('積分計算方式', ''))
        ]
        for info in basic_info:
            ws_detail.cell(row=current_row, column=1, value=info[0])
            cell = ws_detail.cell(row=current_row, column=2, value=info[1])
            if info[0] == "AI初審詳細說明":
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1

        # 寫入演講主題表格題
        headers = ["主題", "講者", "主持人", "時間", "持續時間(分鐘)", "AI初審"]
        for col, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=current_row, column=col, value=header)
            cell.font = title_font
            cell.alignment = title_alignment
        current_row += 1

        # 寫入演講主題詳情
        for topic in result.get('演講主題', []):
            ws_detail.cell(row=current_row, column=1, value=topic.get('topic', ''))
            ws_detail.cell(row=current_row, column=2, value=topic.get('speaker', ''))
            ws_detail.cell(row=current_row, column=3, value=topic.get('moderator', ''))
            ws_detail.cell(row=current_row, column=4, value=topic.get('time', ''))
            ws_detail.cell(row=current_row, column=5, value=topic.get('duration', 0))
            ws_detail.cell(row=current_row, column=6, value=topic.get('ai_review', ''))
            current_row += 1

        # 添加空行作為分隔
        current_row += 2

    # 特別處理 'AI初審詳細說明' 列
    ws_detail.column_dimensions['A'].width = 80  # 設置固定寬度
    ws_detail.column_dimensions['B'].width = 135  # 設置固定寬度
    ws_detail.column_dimensions['C'].width = 36
    ws_detail.column_dimensions['D'].width = 14
    ws_detail.column_dimensions['E'].width = 15
    ws_detail.column_dimensions['F'].width = 40  # 增加 AI初審 欄位的寬度

    wb.save(output_file)

def is_special_item(topic):
    special_keywords = [
        'registration', '報到', '簽到',
        'opening', '開幕', '開場',
        'closing', '閉幕', '結束',
        'panel discussion', '座談會', '討論會',
        'break', '休息',
        'lunch', '午餐',
        'dinner', '晚餐',
        'welcome', '歡迎'
    ]
    topic_lower = topic.lower()
    return any(keyword in topic_lower for keyword in special_keywords)

def determine_credit_type(organizer):
    original_organizer = organizer
    organizer = organizer.lower().strip()
    
    logging.info(f"開始判斷積分類別 - 原始主辦單位: '{original_organizer}', 處理後: '{organizer}'")
    
    # 定義甲類積分的主單位列表
    class_a_organizers = [
        "中華民國糖尿病學會",
        "糖尿病學會",
        "中華民國內分泌學會",
        "內分泌學會"
    ]
    
    # 定義特定的乙類積分主辦單位
    class_b_organizers = [
        "台灣基層糖尿病協會",
        "基層糖尿病協會"
        "衛教學會"
        "糖尿病衛教學會"
    ]
    
    # 檢查是否為特定的乙類積分主辦單位
    for org in class_b_organizers:
        if org.lower() in organizer:
            logging.info(f"部分匹配到特定乙類主辦單位: '{org}'")
            return "乙類"
    
    # 檢查是否為甲類積分主辦單位
    for org in class_a_organizers:
        if org.lower() == organizer:
            logging.info(f"完全匹配到甲類主辦單位: '{org}'")
            return "甲類"
    
    logging.info("未匹配到特定主辦單位，默認為乙類")
    return "乙類"

def process_single_file(client, file_path, model="gpt-4.1-mini"):
    try:
        analyzed_content = analyze_file(client, file_path, model)
        if not analyzed_content:
            return None

        prompt = f"""請仔細分析以下醫學會議文件內容，並提取指定的資訊。

需要提取的資訊：
1. 主題：會議的完整主題名稱
2. 主辦單位：主辦單位的完整名稱
3. 日期：會議舉辦日期
4. 地點：會議舉辦地點
5. 演講主題：包含以下詳細資訊
   - 主題名稱（包含時間資訊）
   - 講者姓名
   - 主持人姓名
   - 時間（格式：HH:MM-HH:MM）
   - 持續時間（分鐘）
   - AI 初審結果

重要規則：
1. 時間格式必須統一為 HH:MM
2. QA 或問答時間必須合併到對應主題中，不單獨列出
3. 如果主題包含 QA，在主題名稱後加註「（包含 QA）」
4. 持續時間必須包含主題演講時間和 QA 時間的總和
5. 特殊項目（如報到、開幕等）需要包含但不計入積分時間
6. 講者姓名若無法確認請留空，不要猜測
7. AI 初審標準：
   - 相關：與糖尿病、高血壓、高血脂或其併發症相關
   - 不相關：與上述疾病無關（需註明原因）
   - ？：無法確定相關性

請以下列 JSON 格式返回分析結果：
{{
    "主題": "string",
    "主辦單位": "string",
    "日期": "string",
    "地點": "string",
    "積分類別": "string",
    "演講主題": [
        {{
            "topic": "string",
            "speaker": "string",
            "moderator": "string",
            "time": "string",
            "duration": number,
            "ai_review": "string"
        }}
    ]
}}

以下是需要分析的內容：
{analyzed_content}"""

        parsed_result = get_gpt4_json_response(client, prompt, model)
        if parsed_result:
            logging.info(f"OpenAI 解析結果: {json.dumps(parsed_result, ensure_ascii=False, indent=2)}")
            organizer = parsed_result.get('主辦單位', '')
            logging.info(f"從 OpenAI 解析結果中獲取的主辦單位: '{organizer}'")
            
            credit_type = determine_credit_type(organizer)
            parsed_result['積分類別'] = credit_type

            logging.info(f"最終判斷結果 - 主辦單位: '{organizer}', 積分類別: {credit_type}")
        else:
            logging.error("OpenAI 分析失敗")
            return None

        # 確保 '演講題' 字段存在且為列表
        if '演講主題' not in parsed_result or not isinstance(parsed_result['演講主題'], list):
            logger.warning(f"警告：分析結果中沒有有效的 '演講主題' 字段")
            parsed_result['演講主題'] = []

        # 處理主題，重新計算包含 QA 的 duration
        parsed_result['演講主題'] = process_topics(parsed_result['演講主題'])

        # 計算總時間和有效時間
        total_duration = 0
        valid_duration = 0
        ai_review_details = []
        uncertain_items = []
        for topic in parsed_result['演講主題']:
            duration = topic.get('duration', 0)
            total_duration += duration
            
            topic_name = topic.get('topic', '')
            ai_review = topic.get('ai_review', '').lower()
            
            if is_special_item(topic_name):
                ai_review_details.append(f"{topic_name}: 特殊項目，不計入學分計算 (0 分鐘)")
            elif ai_review == '相關':
                valid_duration += duration
                ai_review_details.append(f"{topic_name}: AI 判定相關，計入 ({duration} 分鐘)")
            elif '?' in ai_review:
                uncertain_items.append(f"{topic_name}: AI 判定不確定，不計入 ({duration} 分鐘)")
                ai_review_details.append(f"{topic_name}: AI 判定不確定，不計入 (0 分鐘)")
            else:
                ai_review_details.append(f"{topic_name}: AI 判定不相關，不計入 (0 分鐘)")

        # 計算積分
        credits = calculate_credits(valid_duration, credit_type)

        # 添加計算結果到 parsed_result
        parsed_result['原始積分數'] = calculate_credits(total_duration, credit_type)
        parsed_result['AI初審積分'] = credits
        parsed_result['AI初審積分說明'] = f"有效時間：{valid_duration} 分鐘，總時間：{total_duration} 分鐘"
        
        uncertain_explanation = "需要人工審核的項目：\n" + "\n".join(uncertain_items) if uncertain_items else ""
        
        parsed_result['AI初審詳細說明'] = textwrap.dedent(f"""\
            AI 初審積分計算詳情：
            總時間：{total_duration} 分鐘
            有效時間：{valid_duration} 分鐘
            {credit_type}積分計算方式：{'每 50-60 分鐘 1 分' if credit_type == '甲類' else '每 50-60 分鐘 0.5 分'}
            各主題審查結果：
            {chr(10).join(ai_review_details)}
            
            {uncertain_explanation}
            """).strip()
        parsed_result['積分計算方式'] = (
            f"{credit_type}積分：{'每 50-60 分鐘 1 分' if credit_type == '甲類' else '每 50-60 分鐘 0.5 分'}"
        )

        logging.info(f"AI 初審積分計算詳情：\n{parsed_result['AI初審詳細說明']}")
        logging.info(f"積分計算方式：{parsed_result['積分計算方式']}")
        logging.info(f"最終 AI 初審積分：{credits}")

        return parsed_result
    except Exception as e:
        logger.error(f"處理檔案時發生異常: {str(e)}")
        return None

def main():
    st.title("糖尿病學會 學分分析助手")

    # 添加文件格式說明
    st.markdown("""
    **注意：** 
    - 支援的文件格式：PDF, DOCX, JPG, JPEG, PNG
    - 不支援 .doc 格式，請先將 .doc 文件轉換為 .docx 格式後再上傳
    - 轉換方法：使用 Microsoft Word 打開 .doc 文件，然後「另存新檔」為 .docx 格式
    - 尚不支援複雜PDF文件的處理
    """)

    # 在側邊欄中設置 API 金鑰輸入
    st.sidebar.title("設定")
    openai_api_key = st.sidebar.text_input(
        label="請輸入您的 OpenAI API 金鑰：",
        type='password',
        placeholder="例如：sk-2twmA88un4...",
        help="您可以從 https://platform.openai.com/account/api-keys/ 獲取您的 API 金鑰"
    )
    
    # 添加模型選擇下拉選單
    model_option = st.sidebar.selectbox(
        "選擇使用的模型",
        ["gpt-4.1-mini", "gpt-4o", "o3-mini", "o4-mini"],
        index=0,
        help="選擇要使用的 OpenAI 模型，不同模型有不同的效能和成本"
    )

    # 添加空白空間，將聯絡信息推到底部
    st.sidebar.empty()
    st.sidebar.empty()
    st.sidebar.empty()
    
    # 在側邊欄底部添加聯絡信息
    st.sidebar.markdown("---")  # 添加分隔線
    st.sidebar.markdown("**程式設計:** Tseng Yao Hsien \n Tung's Taichung Metroharbor Hospital")
    st.sidebar.markdown("**聯絡:** LINE: zinojeng")
    st.sidebar.markdown(f"**版本：** Version {VERSION}")

    # 檢查是否輸入了 API 金鑰
    if not openai_api_key:
        st.warning("請在側邊欄輸入有效的 OpenAI API 金鑰以繼續")
        return

    # 使用 API 金鑰，但不存儲它
    client = openai.OpenAI(api_key=openai_api_key)
    
    # 將選擇的模型傳遞給處理函數
    selected_model = model_option

    # 檔案上傳
    uploaded_files = st.file_uploader("上傳檔案（支援 PDF, DOCX, JPG, JPEG, PNG）", accept_multiple_files=True, type=['pdf', 'docx', 'jpg', 'jpeg', 'png'])

    if uploaded_files:
        all_results = []
        progress_bar = st.progress(0)

        for index, uploaded_file in enumerate(uploaded_files):
            original_filename = uploaded_file.name
            st.write(f"正在處理檔案：{original_filename}")

            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1]) as temp_file:
                temp_file.write(uploaded_file.getvalue())
                temp_file_path = temp_file.name

            try:
                # 恢復原始的處理方法
                parsed_result = process_single_file(client, temp_file_path, selected_model)
                
                if parsed_result:
                    parsed_result['文件名'] = original_filename  # 使用原始檔案名
                    all_results.append(parsed_result)
                    st.success(f"成功處理檔案：{original_filename}")
                    
                    # 檢查需要人工審核的項目
                    need_review = [topic for topic in parsed_result['演講主題'] if not topic['speaker'] or topic['ai_review'] == '？']
                    if need_review:
                        st.warning("以下項目可能需要人工審核：")
                        for topic in need_review:
                            st.write(f"- 演講主題: {topic['topic']}, 講者: {'未識別' if not topic['speaker'] else topic['speaker']}")
                else:
                    st.error(f"處理檔案 {original_filename} 時出錯")
            except Exception as e:
                logger.error(f"理檔案 {original_filename} 時發生異常: {str(e)}")
                st.error(f"處理檔案 {original_filename} 時發生異常: {str(e)}")
            finally:
                os.unlink(temp_file_path)
            
            # 更新進度條
            progress_bar.progress((index + 1) / len(uploaded_files))

        if all_results:
            output_file = '分析結果總表.xlsx'
            write_to_excel(all_results, output_file)
            st.success(f"所有結果已輸出到 {output_file}")
            
            with open(output_file, "rb") as file:
                st.download_button(
                    label="下載分析結果",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("沒有成功處理任何檔案")

if __name__ == "__main__":
    main()  # 這會在文件末尾添加一個空行
